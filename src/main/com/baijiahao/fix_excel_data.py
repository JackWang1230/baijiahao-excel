import openpyxl
import datetime
from natsort import natsorted


def read_excel():
    # workbook = openpyxl.load_workbook('/Users/wangrui/Documents/coding/python/baojiahao/resources/123.xlsx')
    workbook = openpyxl.load_workbook(
        '/Users/wangrui/Documents/coding/python/baojiahao/resources/laoyao-say-history.xlsx')
    worksheet = workbook['Sheet']
    n_rows = worksheet.max_row
    start = 1
    index = 0
    for i in worksheet.iter_rows(min_row=2, values_only=True):
        index += 1
        print(index)
        total_read = i[1]
        total_comment = i[3]
        if '万' and '.' in total_read:
            fix_read_value = int(total_read.replace('.', '').replace('万', '') + '000')
            worksheet.cell(index + 1, 2, fix_read_value)
        elif '万' in total_read:
            fix_read_value = int(total_read.replace('万', '') + '0000')
            worksheet.cell(index + 1, 2, fix_read_value)
        else:
            fix_read_value = int(total_read)
            worksheet.cell(index + 1, 2, fix_read_value)

        if '万' and '.' in total_comment:
            fix_comment_value = int(total_comment.replace('.', '').replace('万', '') + '000')
            worksheet.cell(index + 1, 4, fix_comment_value)
        elif '万' in total_comment:
            fix_comment_value = int(total_comment.replace('万', '') + '0000')
            worksheet.cell(index + 1, 4, fix_comment_value)
        else:
            fix_comment_value = int(total_comment)
            worksheet.cell(index + 1, 4, fix_comment_value)

        process = (index - start) / (n_rows - start) * 100
        print("处理进度:", process, "%")
    print("开始保存:", datetime.datetime.now())
    # workbook.save('kanye-fix-total_read.xlsx')
    workbook.save('laoyao-fix-total_read.xlsx')
    print("完成保存:", datetime.datetime.now())


def sort_by_total_read():
    workbook = openpyxl.load_workbook('laoyao-fix-total_read.xlsx')
    # workbook = openpyxl.load_workbook('kanye-fix-total_read.xlsx')
    worksheet = workbook['Sheet']
    sorted_rows = natsorted(worksheet.rows, key=lambda x: x[1].value, reverse=True)
    wb_new = openpyxl.Workbook()
    sheet_new = wb_new.active
    for row in sorted_rows:
        sheet_new.append([cell.value for cell in row])

    print("开始保存:", datetime.datetime.now())
    # wb_new.save('kanye-fixed-total_read.xlsx')
    wb_new.save('laoyao-fixed-total_read.xlsx')
    print("完成保存:", datetime.datetime.now())


if __name__ == '__main__':

    read_excel()
    sort_by_total_read()
