import openpyxl
import os
import requests
from datetime import datetime, timedelta


def read_excel():
    workbook = openpyxl.load_workbook('laoyao-fixed-total_read.xlsx')
    # workbook = openpyxl.load_workbook('kanye-fixed-total_read.xlsx')
    worksheet = workbook['Sheet']
    worksheet.cell(row=1, column=7, value='图片下载路径')
    n_rows = worksheet.max_row
    start = 1
    index = 0
    day_index = 0
    today = datetime.today()
    for i in worksheet.iter_rows(min_row=2, values_only=True):

        if index % 15 == 0:
            folder_time = today + timedelta(days=day_index)
            day_index += 1
            true_day = folder_time.strftime('%Y-%m-%d')
            fold_name = f'{true_day}-start-{index + 2}'
            os.makedirs(fold_name, exist_ok=True)
        index += 1
        url_list = []
        image_url = i[6]
        image_url_list = image_url.split(",https")
        if len(image_url_list) > 1:
            url_list.append(image_url_list[0])
            for v in range(1, len(image_url_list)):
                url_list.append('https' + image_url_list[v])
        else:
            url_list.append(image_url_list[0])

        image_filename = ''
        for j in range(len(url_list)):
            response = requests.get(url_list[j])
            image_filename += str(index + 1) + "-" + str(j) + ".jpeg"
            with open(os.path.join(fold_name, image_filename), 'wb') as f:
                f.write(response.content)

        worksheet.cell(index + 1, 7, image_filename)
        process = (index - start) / (n_rows - start) * 100

        print("处理进度:", process, "%")

    print("开始保存:", datetime.now())
    # workbook.save('kanye_finished.xlsx')
    workbook.save('laoyao_finished.xlsx')
    print("完成保存:", datetime.now())


if __name__ == '__main__':
    read_excel()
